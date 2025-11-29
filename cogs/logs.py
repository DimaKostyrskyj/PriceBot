# -*- coding: utf-8 -*-
import discord
from discord.ext import commands
from datetime import datetime, timedelta
import json
import os
from utils.config_manager import ConfigManager
from utils.permissions import permissions
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class EnhancedLogs(commands.Cog):
    """Улучшенная система логирования с максимальной детализацией"""
    
    def __init__(self, bot):
        self.bot = bot
        self.config = ConfigManager()
        self.logs_dir = "logs"
        
        # Файлы логов
        self.bot_logs_file = os.path.join(self.logs_dir, "bot_logs.json")
        self.discord_logs_file = os.path.join(self.logs_dir, "discord_logs.json")
        self.applications_file = os.path.join(self.logs_dir, "applications.json")
        self.voice_sessions_file = os.path.join(self.logs_dir, "voice_sessions.json")
        
        # Создаем папку
        if not os.path.exists(self.logs_dir):
            os.makedirs(self.logs_dir)
        
        # Создаем файлы
        for file in [self.bot_logs_file, self.discord_logs_file, self.applications_file, self.voice_sessions_file]:
            if not os.path.exists(file):
                with open(file, 'w', encoding='utf-8') as f:
                    json.dump([], f)
        
        # Трекинг войс сессий
        self.voice_sessions = {}
    
    def add_log(self, log_file: str, log_type: str, data: dict, user=None, target=None, channel=None, executor=None):
        """
        Добавление ПОЛНОЙ записи в лог
        user - кто выполняет действие
        target - на ком выполняется действие (для модерации)
        executor - кто модерирует (для банов и т.д.)
        """
        try:
            with open(log_file, 'r', encoding='utf-8') as f:
                logs = json.load(f)
            
            log_entry = {
                'id': len(logs) + 1,
                'timestamp': datetime.now().isoformat(),
                'date': datetime.now().strftime('%d.%m.%Y'),
                'time': datetime.now().strftime('%H:%M:%S'),
                'weekday': datetime.now().strftime('%A'),
                'type': log_type,
                
                # Пользователь (кто делает)
                'user_id': user.id if user else None,
                'user_name': str(user) if user else None,
                'user_nick': user.nick if hasattr(user, 'nick') else None,
                'user_discriminator': user.discriminator if user else None,
                
                # Цель (на ком делают)
                'target_id': target.id if target else None,
                'target_name': str(target) if target else None,
                'target_nick': target.nick if hasattr(target, 'nick') else None,
                
                # Модератор (кто модерирует)
                'executor_id': executor.id if executor else None,
                'executor_name': str(executor) if executor else None,
                
                # Канал
                'channel_id': channel.id if channel else None,
                'channel_name': str(channel) if channel else None,
                'channel_type': str(channel.type) if channel else None,
                
                # Данные
                'data': data
            }
            logs.append(log_entry)
            
            if len(logs) > 50000:
                logs = logs[-50000:]
            
            with open(log_file, 'w', encoding='utf-8') as f:
                json.dump(logs, f, indent=2, ensure_ascii=False)
        
        except Exception as e:
            print(f"❌ Ошибка при добавлении лога: {e}")
    
    # ============================================================
    # ГОЛОСОВЫЕ КАНАЛЫ С ТРЕКИНГОМ ВРЕМЕНИ
    # ============================================================
    
    @commands.Cog.listener()
    async def on_voice_state_update(self, member: discord.Member, before: discord.VoiceState, after: discord.VoiceState):
        """Полное логирование войса с временем"""
        
        # Вход в войс
        if before.channel is None and after.channel is not None:
            # Начинаем сессию
            session_id = f"{member.id}_{datetime.now().timestamp()}"
            self.voice_sessions[member.id] = {
                'session_id': session_id,
                'start_time': datetime.now(),
                'channel': after.channel.name,
                'channel_id': after.channel.id
            }
            
            self.add_log(self.discord_logs_file, 'voice_join', {
                'channel': after.channel.name,
                'channel_id': after.channel.id,
                'session_id': session_id,
                'members_in_channel': len(after.channel.members)
            }, user=member, channel=after.channel)
            
            await self._send_log_embed(
                title='🎤 Вход в голосовой канал',
                description=(
                    f"**Пользователь:** {member.mention}\n"
                    f"**Канал:** {after.channel.mention}\n"
                    f"**Участников в канале:** {len(after.channel.members)}"
                ),
                fields=[
                    {'name': 'ID пользователя', 'value': f'`{member.id}`', 'inline': True},
                    {'name': 'Время', 'value': datetime.now().strftime('%H:%M:%S'), 'inline': True}
                ],
                color=0x43B581
            )
        
        # Выход из войса
        elif before.channel is not None and after.channel is None:
            # Завершаем сессию
            session = self.voice_sessions.get(member.id)
            duration = None
            
            if session:
                duration_seconds = (datetime.now() - session['start_time']).total_seconds()
                duration = self._format_duration(duration_seconds)
                del self.voice_sessions[member.id]
                
                # Сохраняем сессию
                self._save_voice_session(member, session, duration_seconds)
            
            self.add_log(self.discord_logs_file, 'voice_leave', {
                'channel': before.channel.name,
                'channel_id': before.channel.id,
                'duration': duration,
                'members_left_in_channel': len(before.channel.members) - 1
            }, user=member, channel=before.channel)
            
            await self._send_log_embed(
                title='🎤 Выход из голосового канала',
                description=(
                    f"**Пользователь:** {member.mention}\n"
                    f"**Канал:** {before.channel.mention}\n"
                    f"**Провел в войсе:** {duration or 'Неизвестно'}"
                ),
                fields=[
                    {'name': 'ID пользователя', 'value': f'`{member.id}`', 'inline': True},
                    {'name': 'Время', 'value': datetime.now().strftime('%H:%M:%S'), 'inline': True}
                ],
                color=0xF04747
            )
        
        # Перемещение между войсами
        elif before.channel != after.channel and before.channel and after.channel:
            self.add_log(self.discord_logs_file, 'voice_move', {
                'from_channel': before.channel.name,
                'from_channel_id': before.channel.id,
                'to_channel': after.channel.name,
                'to_channel_id': after.channel.id,
                'members_in_old': len(before.channel.members) - 1,
                'members_in_new': len(after.channel.members)
            }, user=member)
            
            await self._send_log_embed(
                title='🔀 Перемещение в голосовом канале',
                description=f"**Пользователь:** {member.mention}",
                fields=[
                    {'name': '📤 Из канала', 'value': f'{before.channel.mention}\n({len(before.channel.members) - 1} осталось)', 'inline': True},
                    {'name': '📥 В канал', 'value': f'{after.channel.mention}\n({len(after.channel.members)} участников)', 'inline': True},
                    {'name': 'ID', 'value': f'`{member.id}`', 'inline': False}
                ],
                color=0xFAA61A
            )
        
        # Отключение/включение микрофона
        if before.self_mute != after.self_mute:
            status = "🔇 Выключил" if after.self_mute else "🎤 Включил"
            self.add_log(self.discord_logs_file, 'voice_mute_toggle', {
                'muted': after.self_mute,
                'channel': after.channel.name if after.channel else None
            }, user=member, channel=after.channel)
            
            if after.channel:
                await self._send_log_embed(
                    title=f'{status} микрофон',
                    description=(
                        f"**Пользователь:** {member.mention}\n"
                        f"**Канал:** {after.channel.mention}"
                    ),
                    fields=[],
                    color=0x5865F2
                )
        
        # Отключение/включение звука
        if before.self_deaf != after.self_deaf:
            status = "🔇 Выключил" if after.self_deaf else "🔊 Включил"
            self.add_log(self.discord_logs_file, 'voice_deaf_toggle', {
                'deafened': after.self_deaf,
                'channel': after.channel.name if after.channel else None
            }, user=member, channel=after.channel)
        
        # Видео
        if before.self_video != after.self_video:
            status = "📹 Включил" if after.self_video else "📴 Выключил"
            self.add_log(self.discord_logs_file, 'voice_video_toggle', {
                'video_enabled': after.self_video,
                'channel': after.channel.name if after.channel else None
            }, user=member, channel=after.channel)
            
            if after.channel:
                await self._send_log_embed(
                    title=f'{status} видео',
                    description=(
                        f"**Пользователь:** {member.mention}\n"
                        f"**Канал:** {after.channel.mention}"
                    ),
                    fields=[],
                    color=0x5865F2
                )
        
        # Стрим
        if before.self_stream != after.self_stream:
            status = "🔴 Начал" if after.self_stream else "⚫ Закончил"
            self.add_log(self.discord_logs_file, 'voice_stream_toggle', {
                'streaming': after.self_stream,
                'channel': after.channel.name if after.channel else None
            }, user=member, channel=after.channel)
            
            if after.channel:
                await self._send_log_embed(
                    title=f'{status} стрим',
                    description=(
                        f"**Пользователь:** {member.mention}\n"
                        f"**Канал:** {after.channel.mention}"
                    ),
                    fields=[],
                    color=0xFF0000 if after.self_stream else 0x5865F2
                )
    
    def _format_duration(self, seconds):
        """Форматирование длительности"""
        hours = int(seconds // 3600)
        minutes = int((seconds % 3600) // 60)
        secs = int(seconds % 60)
        
        if hours > 0:
            return f"{hours}ч {minutes}м {secs}с"
        elif minutes > 0:
            return f"{minutes}м {secs}с"
        else:
            return f"{secs}с"
    
    def _save_voice_session(self, member, session, duration_seconds):
        """Сохранение войс сессии"""
        try:
            with open(self.voice_sessions_file, 'r', encoding='utf-8') as f:
                sessions = json.load(f)
            
            sessions.append({
                'user_id': member.id,
                'user_name': str(member),
                'channel': session['channel'],
                'start_time': session['start_time'].isoformat(),
                'end_time': datetime.now().isoformat(),
                'duration_seconds': duration_seconds,
                'duration_formatted': self._format_duration(duration_seconds)
            })
            
            with open(self.voice_sessions_file, 'w', encoding='utf-8') as f:
                json.dump(sessions, f, indent=2, ensure_ascii=False)
        except Exception as e:
            print(f"❌ Ошибка сохранения сессии: {e}")
    
    # ============================================================
    # ПРИГЛАШЕНИЯ
    # ============================================================
    
    @commands.Cog.listener()
    async def on_invite_create(self, invite: discord.Invite):
        """Создание приглашений"""
        self.add_log(self.discord_logs_file, 'invite_create', {
            'code': invite.code,
            'url': invite.url,
            'channel': str(invite.channel),
            'max_uses': invite.max_uses or 'Unlimited',
            'max_age': invite.max_age or 'Never',
            'temporary': invite.temporary
        }, user=invite.inviter)
        
        await self._send_log_embed(
            title='🔗 Создано приглашение',
            description=f"**Создал:** {invite.inviter.mention if invite.inviter else 'Неизвестно'}",
            fields=[
                {'name': 'Код', 'value': f'`{invite.code}`', 'inline': True},
                {'name': 'Канал', 'value': invite.channel.mention if hasattr(invite.channel, 'mention') else str(invite.channel), 'inline': True},
                {'name': 'Макс. использований', 'value': str(invite.max_uses or '∞'), 'inline': True},
                {'name': 'Макс. время', 'value': self._format_duration(invite.max_age) if invite.max_age else '∞', 'inline': True},
                {'name': 'URL', 'value': f'`{invite.url}`', 'inline': False}
            ],
            color=0x43B581
        )
    
    @commands.Cog.listener()
    async def on_invite_delete(self, invite: discord.Invite):
        """Удаление приглашений"""
        self.add_log(self.discord_logs_file, 'invite_delete', {
            'code': invite.code,
            'channel': str(invite.channel),
            'uses': invite.uses,
            'max_uses': invite.max_uses
        })
        
        await self._send_log_embed(
            title='🔗 Удалено приглашение',
            description=f"**Код:** `{invite.code}`",
            fields=[
                {'name': 'Использовано', 'value': f'{invite.uses}/{invite.max_uses or "∞"}', 'inline': True}
            ],
            color=0xF04747
        )
    
    # ============================================================
    # СООБЩЕНИЯ
    # ============================================================
    
    @commands.Cog.listener()
    async def on_message_delete(self, message: discord.Message):
        """Удаление сообщений"""
        if message.author.bot:
            return
        
        self.add_log(self.discord_logs_file, 'message_delete', {
            'content': message.content[:500],
            'content_length': len(message.content),
            'attachments': [att.filename for att in message.attachments],
            'embeds_count': len(message.embeds),
            'mentions': [str(m) for m in message.mentions[:5]],
            'message_id': message.id
        }, user=message.author, channel=message.channel)
        
        embed_desc = (
            f"**Автор:** {message.author.mention}\n"
            f"**Канал:** {message.channel.mention}\n"
            f"**ID сообщения:** `{message.id}`"
        )
        
        fields = [
            {'name': 'Содержание', 'value': message.content[:1000] if message.content else '*Пусто*', 'inline': False}
        ]
        
        if message.attachments:
            fields.append({
                'name': f'Вложения ({len(message.attachments)})',
                'value': '\n'.join([f'`{att.filename}`' for att in message.attachments[:5]]),
                'inline': False
            })
        
        await self._send_log_embed(
            title='🗑️ Сообщение удалено',
            description=embed_desc,
            fields=fields,
            color=0xF04747
        )
    
    @commands.Cog.listener()
    async def on_message_edit(self, before: discord.Message, after: discord.Message):
        """Редактирование сообщений"""
        if before.author.bot or before.content == after.content:
            return
        
        self.add_log(self.discord_logs_file, 'message_edit', {
            'before': before.content[:500],
            'after': after.content[:500],
            'message_id': after.id,
            'jump_url': after.jump_url
        }, user=before.author, channel=before.channel)
        
        await self._send_log_embed(
            title='✏️ Сообщение отредактировано',
            description=(
                f"**Автор:** {before.author.mention}\n"
                f"**Канал:** {before.channel.mention}\n"
                f"**[Перейти к сообщению]({after.jump_url})**"
            ),
            fields=[
                {'name': '📝 До', 'value': before.content[:500] if before.content else '*Пусто*', 'inline': False},
                {'name': '📝 После', 'value': after.content[:500] if after.content else '*Пусто*', 'inline': False}
            ],
            color=0x5865F2
        )
    
    # ============================================================
    # МОДЕРАЦИЯ (КТО НА КОМ)
    # ============================================================
    
    @commands.Cog.listener()
    async def on_member_ban(self, guild: discord.Guild, user: discord.User):
        """Бан пользователя"""
        # Пытаемся найти кто забанил через audit log
        executor = None
        reason = None
        
        try:
            async for entry in guild.audit_logs(limit=5, action=discord.AuditLogAction.ban):
                if entry.target.id == user.id:
                    executor = entry.user
                    reason = entry.reason
                    break
        except:
            pass
        
        self.add_log(self.discord_logs_file, 'member_ban', {
            'reason': reason or 'Не указана',
            'guild_name': guild.name
        }, user=user, executor=executor)
        
        await self._send_log_embed(
            title='🔨 Пользователь забанен',
            description=(
                f"**Забанен:** {user.mention} (`{user.id}`)\n"
                f"**Модератор:** {executor.mention if executor else 'Неизвестно'}\n"
                f"**Причина:** {reason or 'Не указана'}"
            ),
            fields=[],
            color=0xF04747
        )
    
    @commands.Cog.listener()
    async def on_member_unban(self, guild: discord.Guild, user: discord.User):
        """Разбан пользователя"""
        executor = None
        
        try:
            async for entry in guild.audit_logs(limit=5, action=discord.AuditLogAction.unban):
                if entry.target.id == user.id:
                    executor = entry.user
                    break
        except:
            pass
        
        self.add_log(self.discord_logs_file, 'member_unban', {
            'guild_name': guild.name
        }, user=user, executor=executor)
        
        await self._send_log_embed(
            title='✅ Пользователь разбанен',
            description=(
                f"**Разбанен:** {user.mention} (`{user.id}`)\n"
                f"**Модератор:** {executor.mention if executor else 'Неизвестно'}"
            ),
            fields=[],
            color=0x43B581
        )
    
    @commands.Cog.listener()
    async def on_member_kick(self, guild: discord.Guild, user: discord.User):
        """Кик пользователя"""
        executor = None
        reason = None
        
        try:
            async for entry in guild.audit_logs(limit=5, action=discord.AuditLogAction.kick):
                if entry.target.id == user.id:
                    executor = entry.user
                    reason = entry.reason
                    break
        except:
            pass
        
        self.add_log(self.discord_logs_file, 'member_kick', {
            'reason': reason or 'Не указана',
            'guild_name': guild.name
        }, user=user, executor=executor)
        
        await self._send_log_embed(
            title='👢 Пользователь кикнут',
            description=(
                f"**Кикнут:** {user.mention} (`{user.id}`)\n"
                f"**Модератор:** {executor.mention if executor else 'Неизвестно'}\n"
                f"**Причина:** {reason or 'Не указана'}"
            ),
            fields=[],
            color=0xFAA61A
        )
    
    # ============================================================
    # ПОЛЬЗОВАТЕЛИ
    # ============================================================
    
    @commands.Cog.listener()
    async def on_member_join(self, member: discord.Member):
        """Вход участника"""
        account_age = (datetime.now() - member.created_at.replace(tzinfo=None)).days
        
        self.add_log(self.discord_logs_file, 'member_join', {
            'account_created': member.created_at.isoformat(),
            'account_age_days': account_age,
            'is_bot': member.bot,
            'avatar_url': str(member.display_avatar.url)
        }, user=member)
        
        await self._send_log_embed(
            title='👋 Участник присоединился',
            description=f"**{member.mention}**",
            fields=[
                {'name': 'ID', 'value': f'`{member.id}`', 'inline': True},
                {'name': 'Аккаунт создан', 'value': f'{account_age} дн. назад', 'inline': True},
                {'name': 'Тип', 'value': '🤖 Бот' if member.bot else '👤 Пользователь', 'inline': True}
            ],
            color=0x43B581,
            thumbnail=member.display_avatar.url
        )
    
    @commands.Cog.listener()
    async def on_member_remove(self, member: discord.Member):
        """Выход участника"""
        roles = [role.name for role in member.roles if role.name != "@everyone"]
        
        self.add_log(self.discord_logs_file, 'member_leave', {
            'roles': roles,
            'joined_at': member.joined_at.isoformat() if member.joined_at else None,
            'nickname': member.nick,
            'roles_count': len(roles)
        }, user=member)
        
        await self._send_log_embed(
            title='👋 Участник покинул сервер',
            description=f"**{member.mention}**",
            fields=[
                {'name': 'ID', 'value': f'`{member.id}`', 'inline': True},
                {'name': 'Ролей', 'value': str(len(roles)), 'inline': True},
                {'name': 'Роли', 'value': ', '.join(roles[:10]) if roles else 'Нет', 'inline': False}
            ],
            color=0xF04747,
            thumbnail=member.display_avatar.url
        )
    
    @commands.Cog.listener()
    async def on_member_update(self, before: discord.Member, after: discord.Member):
        """Изменения участника"""
        
        # Изменение ника
        if before.nick != after.nick:
            # Проверяем кто изменил через audit log
            executor = None
            try:
                async for entry in after.guild.audit_logs(limit=5, action=discord.AuditLogAction.member_update):
                    if entry.target.id == after.id:
                        executor = entry.user
                        break
            except:
                pass
            
            self.add_log(self.discord_logs_file, 'member_nick_change', {
                'before': before.nick or before.name,
                'after': after.nick or after.name
            }, user=after, executor=executor)
            
            await self._send_log_embed(
                title='📝 Изменение ника',
                description=(
                    f"**Пользователь:** {after.mention}\n"
                    f"**Изменил:** {executor.mention if executor and executor != after else 'Сам'}"
                ),
                fields=[
                    {'name': '📝 До', 'value': before.nick or before.name, 'inline': True},
                    {'name': '📝 После', 'value': after.nick or after.name, 'inline': True}
                ],
                color=0x5865F2
            )
        
        # Изменение ролей
        before_roles = set(before.roles)
        after_roles = set(after.roles)
        
        added_roles = after_roles - before_roles
        removed_roles = before_roles - after_roles
        
        # Кто добавил/удалил роли
        executor = None
        try:
            async for entry in after.guild.audit_logs(limit=5, action=discord.AuditLogAction.member_role_update):
                if entry.target.id == after.id:
                    executor = entry.user
                    break
        except:
            pass
        
        # Добавление ролей
        for role in added_roles:
            if role.name != "@everyone":
                self.add_log(self.discord_logs_file, 'member_role_add', {
                    'role_name': role.name,
                    'role_id': role.id,
                    'role_color': str(role.color)
                }, user=after, executor=executor)
                
                await self._send_log_embed(
                    title='🎭 Роль добавлена',
                    description=(
                        f"**Пользователь:** {after.mention}\n"
                        f"**Роль:** {role.mention}\n"
                        f"**Модератор:** {executor.mention if executor else 'Автоматически'}"
                    ),
                    fields=[
                        {'name': 'ID пользователя', 'value': f'`{after.id}`', 'inline': True},
                        {'name': 'ID роли', 'value': f'`{role.id}`', 'inline': True}
                    ],
                    color=0x43B581
                )
        
        # Удаление ролей
        for role in removed_roles:
            if role.name != "@everyone":
                self.add_log(self.discord_logs_file, 'member_role_remove', {
                    'role_name': role.name,
                    'role_id': role.id
                }, user=after, executor=executor)
                
                await self._send_log_embed(
                    title='🎭 Роль удалена',
                    description=(
                        f"**Пользователь:** {after.mention}\n"
                        f"**Роль:** {role.name}\n"
                        f"**Модератор:** {executor.mention if executor else 'Автоматически'}"
                    ),
                    fields=[
                        {'name': 'ID пользователя', 'value': f'`{after.id}`', 'inline': True},
                        {'name': 'ID роли', 'value': f'`{role.id}`', 'inline': True}
                    ],
                    color=0xF04747
                )
    
    # ============================================================
    # КАНАЛЫ
    # ============================================================
    
    @commands.Cog.listener()
    async def on_guild_channel_create(self, channel: discord.abc.GuildChannel):
        """Создание канала"""
        executor = None
        try:
            async for entry in channel.guild.audit_logs(limit=5, action=discord.AuditLogAction.channel_create):
                if entry.target.id == channel.id:
                    executor = entry.user
                    break
        except:
            pass
        
        self.add_log(self.discord_logs_file, 'channel_create', {
            'channel_name': channel.name,
            'channel_type': str(channel.type),
            'category': channel.category.name if channel.category else None
        }, executor=executor)
        
        await self._send_log_embed(
            title='➕ Канал создан',
            description=(
                f"**Канал:** {channel.mention}\n"
                f"**Создал:** {executor.mention if executor else 'Неизвестно'}"
            ),
            fields=[
                {'name': 'Тип', 'value': str(channel.type), 'inline': True},
                {'name': 'ID', 'value': f'`{channel.id}`', 'inline': True}
            ],
            color=0x43B581
        )
    
    @commands.Cog.listener()
    async def on_guild_channel_delete(self, channel: discord.abc.GuildChannel):
        """Удаление канала"""
        executor = None
        try:
            async for entry in channel.guild.audit_logs(limit=5, action=discord.AuditLogAction.channel_delete):
                if entry.target.id == channel.id:
                    executor = entry.user
                    break
        except:
            pass
        
        self.add_log(self.discord_logs_file, 'channel_delete', {
            'channel_name': channel.name,
            'channel_type': str(channel.type)
        }, executor=executor)
        
        await self._send_log_embed(
            title='➖ Канал удален',
            description=(
                f"**Канал:** {channel.name}\n"
                f"**Удалил:** {executor.mention if executor else 'Неизвестно'}"
            ),
            fields=[
                {'name': 'Тип', 'value': str(channel.type), 'inline': True},
                {'name': 'ID', 'value': f'`{channel.id}`', 'inline': True}
            ],
            color=0xF04747
        )
    
    # ============================================================
    # РОЛИ
    # ============================================================
    
    @commands.Cog.listener()
    async def on_guild_role_create(self, role: discord.Role):
        """Создание роли"""
        executor = None
        try:
            async for entry in role.guild.audit_logs(limit=5, action=discord.AuditLogAction.role_create):
                if entry.target.id == role.id:
                    executor = entry.user
                    break
        except:
            pass
        
        self.add_log(self.discord_logs_file, 'role_create', {
            'role_name': role.name,
            'role_id': role.id,
            'color': str(role.color),
            'permissions': role.permissions.value
        }, executor=executor)
        
        await self._send_log_embed(
            title='🎭 Роль создана',
            description=(
                f"**Роль:** {role.mention}\n"
                f"**Создал:** {executor.mention if executor else 'Неизвестно'}"
            ),
            fields=[
                {'name': 'ID', 'value': f'`{role.id}`', 'inline': True},
                {'name': 'Цвет', 'value': str(role.color), 'inline': True}
            ],
            color=0x43B581
        )
    
    @commands.Cog.listener()
    async def on_guild_role_delete(self, role: discord.Role):
        """Удаление роли"""
        executor = None
        try:
            async for entry in role.guild.audit_logs(limit=5, action=discord.AuditLogAction.role_delete):
                if entry.target.id == role.id:
                    executor = entry.user
                    break
        except:
            pass
        
        self.add_log(self.discord_logs_file, 'role_delete', {
            'role_name': role.name,
            'role_id': role.id
        }, executor=executor)
        
        await self._send_log_embed(
            title='🎭 Роль удалена',
            description=(
                f"**Роль:** {role.name}\n"
                f"**Удалил:** {executor.mention if executor else 'Неизвестно'}"
            ),
            fields=[
                {'name': 'ID', 'value': f'`{role.id}`', 'inline': True}
            ],
            color=0xF04747
        )
    
    # ============================================================
    # КОМАНДЫ БОТА
    # ============================================================
    
    @commands.Cog.listener()
    async def on_command(self, ctx):
        """Использование команд"""
        self.add_log(self.bot_logs_file, 'command_use', {
            'command': ctx.command.name,
            'full_message': ctx.message.content,
            'args': str(ctx.args[2:]) if len(ctx.args) > 2 else '',
            'success': True
        }, user=ctx.author, channel=ctx.channel)
        
        await self._send_log_embed(
            title='⚡ Команда использована',
            description=(
                f"**Пользователь:** {ctx.author.mention}\n"
                f"**Команда:** `{ctx.command.name}`\n"
                f"**Канал:** {ctx.channel.mention}"
            ),
            fields=[
                {'name': 'Полная команда', 'value': f'`{ctx.message.content}`', 'inline': False}
            ],
            color=0x5865F2
        )
    
    @commands.Cog.listener()
    async def on_command_error(self, ctx, error):
        """Ошибки команд"""
        self.add_log(self.bot_logs_file, 'command_error', {
            'command': ctx.command.name if ctx.command else 'unknown',
            'error': str(error)[:500],
            'error_type': type(error).__name__
        }, user=ctx.author, channel=ctx.channel)
    
    # ============================================================
    # ЗАЯВКИ
    # ============================================================
    
    def log_application(self, user, data: dict):
        """Логирование заявки"""
        self.add_log(self.applications_file, 'application_submit', data, user=user)
    
    def log_application_review(self, executor, applicant, action: str, data: dict):
        """Логирование рассмотрения заявки"""
        self.add_log(self.applications_file, f'application_{action}', data, user=applicant, executor=executor)
    
    # ============================================================
    # СКАЧИВАНИЕ ЛОГОВ
    # ============================================================
    
    @commands.command(name='download_all_logs')
    async def download_all_logs(self, ctx, days: int = 30):
        """Скачать ВСЕ логи в улучшенном Excel (Owner/Developer)"""
        if not permissions.can_use_all_commands(ctx.author):
            await ctx.send('❌ Нет прав! Требуется Owner или Developer.')
            return
        
        await ctx.send('⏳ Создаю улучшенный многостраничный Excel файл...')
        
        try:
            # Читаем логи
            with open(self.bot_logs_file, 'r', encoding='utf-8') as f:
                bot_logs = json.load(f)
            
            with open(self.discord_logs_file, 'r', encoding='utf-8') as f:
                discord_logs = json.load(f)
            
            with open(self.applications_file, 'r', encoding='utf-8') as f:
                applications = json.load(f)
            
            with open(self.voice_sessions_file, 'r', encoding='utf-8') as f:
                voice_sessions = json.load(f)
            
            # Фильтр по дням
            cutoff_date = datetime.now() - timedelta(days=days)
            
            bot_logs_filtered = [l for l in bot_logs if datetime.fromisoformat(l['timestamp']) > cutoff_date]
            discord_logs_filtered = [l for l in discord_logs if datetime.fromisoformat(l['timestamp']) > cutoff_date]
            applications_filtered = [l for l in applications if datetime.fromisoformat(l['timestamp']) > cutoff_date]
            voice_sessions_filtered = [s for s in voice_sessions if datetime.fromisoformat(s['start_time']) > cutoff_date]
            
            # Создаем Excel
            filename = await self._create_enhanced_excel(
                bot_logs_filtered,
                discord_logs_filtered,
                applications_filtered,
                voice_sessions_filtered,
                days
            )
            
            # Отправляем
            embed = discord.Embed(
                title='📥 Полные логи готовы',
                description=f'Улучшенные логи за последние {days} дней',
                color=0x43B581
            )
            embed.add_field(name='📊 Логи бота', value=str(len(bot_logs_filtered)), inline=True)
            embed.add_field(name='💬 Логи Discord', value=str(len(discord_logs_filtered)), inline=True)
            embed.add_field(name='📋 Заявки', value=str(len(applications_filtered)), inline=True)
            embed.add_field(name='🎤 Войс сессии', value=str(len(voice_sessions_filtered)), inline=True)
            embed.add_field(name='📄 Страниц', value='4', inline=True)
            
            await ctx.send(embed=embed, file=discord.File(filename))
            
            os.remove(filename)
            
            self.add_log(self.bot_logs_file, 'logs_download_full', {
                'days': days,
                'bot_logs': len(bot_logs_filtered),
                'discord_logs': len(discord_logs_filtered),
                'applications': len(applications_filtered),
                'voice_sessions': len(voice_sessions_filtered)
            }, user=ctx.author)
        
        except Exception as e:
            await ctx.send(f'❌ Ошибка: {e}')
            print(f"❌ Ошибка: {e}")
            import traceback
            traceback.print_exc()
    
    async def _create_enhanced_excel(self, bot_logs, discord_logs, applications, voice_sessions, days):
        """Создание улучшенного Excel файла"""
        filename = f"full_logs_{days}days.xlsx"
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        # Создаем страницы
        self._create_bot_logs_page(wb, bot_logs)
        self._create_discord_logs_page(wb, discord_logs)
        self._create_applications_page(wb, applications)
        self._create_voice_sessions_page(wb, voice_sessions)
        
        wb.save(filename)
        return filename
    
    def _create_bot_logs_page(self, wb, logs):
        """Страница: Логи бота"""
        ws = wb.create_sheet("📊 Логи Бота")
        
        # Красивая шапка
        header_fill = PatternFill(start_color="5865F2", end_color="5865F2", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # Заголовки
        headers = ['№', 'Дата', 'Время', 'День недели', 'Команда', 'Пользователь', 'ID пользователя', 'Канал', 'Полная команда', 'Результат']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        # Данные
        for idx, log in enumerate(logs, 2):
            ws.cell(row=idx, column=1, value=log.get('id', idx-1)).border = border
            ws.cell(row=idx, column=2, value=log.get('date', '')).border = border
            ws.cell(row=idx, column=3, value=log.get('time', '')).border = border
            ws.cell(row=idx, column=4, value=log.get('weekday', '')).border = border
            ws.cell(row=idx, column=5, value=log.get('type', '')).border = border
            ws.cell(row=idx, column=6, value=log.get('user_name', '')).border = border
            ws.cell(row=idx, column=7, value=log.get('user_id', '')).border = border
            ws.cell(row=idx, column=8, value=log.get('channel_name', '')).border = border
            ws.cell(row=idx, column=9, value=log.get('data', {}).get('full_message', '')).border = border
            ws.cell(row=idx, column=10, value='✅ Успех' if log.get('data', {}).get('success') else '❌ Ошибка').border = border
        
        # Автоширина
        for col in range(1, len(headers) + 1):
            max_length = len(headers[col-1])
            for row in range(2, len(logs) + 2):
                cell_value = str(ws.cell(row=row, column=col).value or '')
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            ws.column_dimensions[get_column_letter(col)].width = min(max_length + 2, 50)
    
    def _create_discord_logs_page(self, wb, logs):
        """Страница: Логи Discord"""
        ws = wb.create_sheet("💬 Логи Discord")
        
        header_fill = PatternFill(start_color="2F3136", end_color="2F3136", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        headers = ['№', 'Дата', 'Время', 'Тип', 'Пользователь (кто)', 'ID', 'Цель (на ком)', 'Модератор', 'Канал', 'Детали']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        for idx, log in enumerate(logs, 2):
            ws.cell(row=idx, column=1, value=log.get('id', idx-1)).border = border
            ws.cell(row=idx, column=2, value=log.get('date', '')).border = border
            ws.cell(row=idx, column=3, value=log.get('time', '')).border = border
            ws.cell(row=idx, column=4, value=log.get('type', '')).border = border
            ws.cell(row=idx, column=5, value=log.get('user_name', '')).border = border
            ws.cell(row=idx, column=6, value=log.get('user_id', '')).border = border
            ws.cell(row=idx, column=7, value=log.get('target_name', '')).border = border
            ws.cell(row=idx, column=8, value=log.get('executor_name', '')).border = border
            ws.cell(row=idx, column=9, value=log.get('channel_name', '')).border = border
            ws.cell(row=idx, column=10, value=json.dumps(log.get('data', {}), ensure_ascii=False)[:200]).border = border
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20
    
    def _create_applications_page(self, wb, applications):
        """Страница: Заявки"""
        ws = wb.create_sheet("📋 Заявки")
        
        header_fill = PatternFill(start_color="43B581", end_color="43B581", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        headers = ['№', 'Дата', 'Время', 'Действие', 'Заявитель', 'ID заявителя', 'Проверяющий', 'Детали']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        for idx, log in enumerate(applications, 2):
            ws.cell(row=idx, column=1, value=log.get('id', idx-1)).border = border
            ws.cell(row=idx, column=2, value=log.get('date', '')).border = border
            ws.cell(row=idx, column=3, value=log.get('time', '')).border = border
            ws.cell(row=idx, column=4, value=log.get('type', '')).border = border
            ws.cell(row=idx, column=5, value=log.get('user_name', '')).border = border
            ws.cell(row=idx, column=6, value=log.get('user_id', '')).border = border
            ws.cell(row=idx, column=7, value=log.get('executor_name', '')).border = border
            ws.cell(row=idx, column=8, value=json.dumps(log.get('data', {}), ensure_ascii=False)[:200]).border = border
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20
    
    def _create_voice_sessions_page(self, wb, sessions):
        """Страница: Голосовые сессии"""
        ws = wb.create_sheet("🎤 Голосовые Сессии")
        
        header_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        headers = ['№', 'Пользователь', 'Канал', 'Начало', 'Конец', 'Длительность', 'Секунд']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        for idx, session in enumerate(sessions, 2):
            ws.cell(row=idx, column=1, value=idx-1).border = border
            ws.cell(row=idx, column=2, value=session.get('user_name', '')).border = border
            ws.cell(row=idx, column=3, value=session.get('channel', '')).border = border
            ws.cell(row=idx, column=4, value=session.get('start_time', '')).border = border
            ws.cell(row=idx, column=5, value=session.get('end_time', '')).border = border
            ws.cell(row=idx, column=6, value=session.get('duration_formatted', '')).border = border
            ws.cell(row=idx, column=7, value=session.get('duration_seconds', 0)).border = border
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 25
    
    async def _send_log_embed(self, title: str, description: str, fields: list, color: int, thumbnail: str = None):
        """Отправка в канал логов"""
        logs_channel_id = self.config.get('logs_channel_id')
        if not logs_channel_id:
            return
        
        logs_channel = self.bot.get_channel(logs_channel_id)
        if not logs_channel:
            return
        
        embed = discord.Embed(
            title=title,
            description=description,
            color=color,
            timestamp=datetime.now()
        )
        
        if thumbnail:
            embed.set_thumbnail(url=thumbnail)
        
        for field in fields:
            embed.add_field(
                name=field['name'],
                value=field['value'],
                inline=field.get('inline', False)
            )
        
        try:
            await logs_channel.send(embed=embed)
        except:
            pass


async def setup(bot):
    await bot.add_cog(EnhancedLogs(bot))